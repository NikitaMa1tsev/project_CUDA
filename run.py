import itertools
from matplotlib import pyplot as plt
from sccf import SccfGPU, SccfCPU
from openpyxl import Workbook
import wmi


def gpu_check() -> bool:
    computer = wmi.WMI()
    gpu_info = computer.Win32_VideoController()[0]
    if gpu_info.Name[0] == "N":
        return True
    else:
        return False


class Test:
    def __init__(self):
        self.num = 1

    def testing(self):
        tau = [1.0, 1.8]
        dts0 = [0, 0.4]
        dt0 = [0, 0.06]
        fs = [10, 20, 40, 60, 80, 100, 250]
        combinations = itertools.product([20.03], tau, [1], dts0, dt0, fs)
        for combination in combinations:
            print(f'testing: {self.num / 56 * 100}%')
            self.run_test(*combination)
            self.num += 1

    def run_test(self, t_imp, tau, f0, dts0, dt0, fs):
        if gpu_check():
            sccf = SccfGPU(t_imp, tau, f0, dts0, dt0, fs)
        else:
            sccf = SccfCPU(t_imp, tau, f0, dts0, dt0, fs)
        sccf.run()
        self.print_plot(sccf)
        self.output_in_excel(sccf)

    def output_in_excel(self, sccf):
        wb = Workbook()
        ws = wb.active
        ws.title = "List_1"
        wb.create_sheet("List_2")
        wb.create_sheet("List_3")

        ws = wb["List_1"]
        ws.cell(1, 1).value = "u(t)"
        for r, statN in enumerate(sccf.u, start=2):
            ws.cell(row=r, column=1).value = statN

        ws.cell(1, 2).value = "w(t)"
        for r, statN in enumerate(sccf.w, start=2):
            ws.cell(row=r, column=2).value = statN

        ws = wb["List_2"]
        for x in list(sccf.c):
            ws.append(list(x))

        ws = wb["List_3"]
        for r, statN in enumerate(sccf.c_sum, start=1):
            ws.cell(row=r, column=1).value = statN

        wb.save('sources/tables/sccf_' + str(self.num) + '.xlsx')

    def print_plot(self, sccf):
        plt.figure(figsize=(25, 12))

        plt.subplot(1, 2, 1)
        plt.plot(sccf.t[::100], sccf.y0[::100], marker='o', markersize=0.5, label='y0(t)', color='black')
        plt.plot(sccf.tu_samples, sccf.u, marker='o', markersize=1, label='u(t)', color='r')
        plt.plot(sccf.tw_samples, sccf.w, marker='o', markersize=1, label='w(t)', color='b')
        plt.legend()
        plt.xlabel("t")
        plt.ylabel("function")

        plt.subplot(1, 2, 2)
        plt.plot(sccf.tu_samples[:len(sccf.c_sum)], sccf.c_sum, marker='o', markersize=1, label='c_sum', color='g')
        plt.legend()
        plt.figtext(0.15, 0.95, f"f0 = {sccf.f0 / 10 ** 6} МГц")
        plt.figtext(0.15, 0.9, f"tau = {(sccf.tau * 10 ** 6):.1f} мкс")
        plt.figtext(0.45, 0.95, f"dts0 = {sccf.dts0 * 10 ** 6} ")
        plt.figtext(0.45, 0.9, f"dt0 = {sccf.dt0 * 10 ** 6} мкс")
        plt.figtext(0.75, 0.95, f"Fs = {sccf.fs / 10 ** 6} МГц")
        plt.xlabel("M")
        plt.ylabel("csum")

        plt.savefig('sources/plots/sccf_' + str(self.num) + '.png')
        plt.close()


class TestOne(Test):
    def print_plot(self, sccf):
        plt.figure(figsize=(25, 12))

        plt.subplot(1, 2, 1)
        plt.plot(sccf.t[::100], sccf.y0[::100], marker='o', markersize=0.5, label='y0(t)', color='black')
        plt.plot(sccf.tu_samples, sccf.u, marker='o', markersize=1, label='u(t)', color='r')
        plt.plot(sccf.tw_samples, sccf.w, marker='o', markersize=1, label='w(t)', color='b')
        plt.legend()
        plt.xlabel("t")
        plt.ylabel("function")

        plt.subplot(1, 2, 2)
        plt.plot(sccf.tu_samples[:len(sccf.c_sum)], sccf.c_sum, marker='o', markersize=1, label='c_sum', color='g')
        plt.legend()
        plt.figtext(0.15, 0.95, f"f0 = {sccf.f0 / 10 ** 6} МГц")
        plt.figtext(0.15, 0.9, f"tau = {(sccf.tau * 10 ** 6):.1f} мкс")
        plt.figtext(0.45, 0.95, f"dts0 = {sccf.dts0 * 10 ** 6} ")
        plt.figtext(0.45, 0.9, f"dt0 = {sccf.dt0 * 10 ** 6} мкс")
        plt.figtext(0.75, 0.95, f"Fs = {sccf.fs / 10 ** 6} МГц")
        plt.xlabel("M")
        plt.ylabel("csum")

        plt.grid()
        plt.show()
